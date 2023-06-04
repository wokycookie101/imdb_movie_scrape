from bs4 import BeautifulSoup
import requests
import openpyxl


def main():
    
    # Opens the excel spreadsheet 
    excel = openpyxl.Workbook()
    print(excel.sheetnames)
    sheet = excel.active
    sheet.title = "Top Rated Foreign Moves"
    print(excel.sheetnames)
    sheet.append(["Rank", "Movie Name", "Year of Release", "Runtime", "Genre", "Official IMDB Rating"])
    
    #scrapes the movies from IMDB
    try:
        # URL I'm using
        source = requests.get("https://www.imdb.com/list/ls052393071/")
        # In case the URL isn't the correct one.
        source.raise_for_status()
        
        soup = BeautifulSoup(source.text, "html.parser")
        movies = soup.find("div", class_="lister-list" ).find_all("div", class_="lister-item-content")
        
        # This code gets individual data from each movie read by the program
        for movie in movies:
            
            name = movie.find("h3", class_="lister-item-header").a.text
            rank = movie.find("span", class_="lister-item-index unbold text-primary").text
            year = movie.find("span", class_="lister-item-year text-muted unbold").text
            runtime = movie.find("span", class_="runtime").text
            genre = movie.find("span", class_="genre").text
            rating = movie.find("span", class_="ipl-rating-star__rating").text
            
            global num_of_dramas
            
            num_of_dramas = 0
            
            #This calculates how many dramas are in this list scraped.
            if genre == "Drama":
                num_of_genres += 1
            else: 
                pass
            
            
            # Printing to check and see if 
            print(rank + " " + year + " " + name)
            print(runtime + " | " + genre)
            print("Rating: " + rating)
            print("\n")
            #Prints this data into an excel sheet.
            sheet.append([rank, name, year, runtime, genre, rating])
            
            
    except Exception:
        print(Exception)
    
    print(f"Number of Dramas: {num_of_dramas}")

    excel.save("IMDB Movie Ratings.xlsx")
    
if __name__ == "__main__":
    main()